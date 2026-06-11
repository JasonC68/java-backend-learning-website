# -*- coding: utf-8 -*-
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# 按复习优先级排序
data = {
    "集合": [
        "HashMap 实现原理", "HashMap 的 put / get 过程", "HashMap 扩容机制",
        "为什么容量是 2 的 n 次方 / 负载因子", "为什么链表转红黑树（而非 AVL）",
        "哈希冲突的解决方法", "HashMap 线程安全吗 / 多线程下的问题",
        "重写 equals / hashCode 的注意点", "HashMap 用什么做 Key / 为什么 String 合适",
        "HashMap vs Hashtable vs ConcurrentHashMap", "ConcurrentHashMap 怎么实现的",
        "用了 synchronized 为什么还要 CAS", "ArrayList vs LinkedList",
        "ArrayList 扩容机制", "ArrayList 为什么线程不安全 / 怎么变安全", "List vs Set",
    ],
    "并发/多线程": [
        "JMM（Java 内存模型）+ 三大问题", "线程的创建方式", "线程的六种状态",
        "sleep vs wait", "notify vs notifyAll", "怎么保证多线程安全",
        "volatile 作用与原理", "synchronized 原理 + 锁升级", "CAS 原理 + ABA 问题",
        "ReentrantLock vs synchronized + AQS", "线程池原理（核心参数/流程/拒绝策略）",
        "BLOCKED vs WAITING / 如何停止线程",
    ],
    "MySQL": [
        "索引为什么用 B+ 树（vs B树/跳表/哈希）", "聚簇索引 vs 非聚簇索引",
        "联合索引 + 最左匹配原则", "索引失效的场景", "回表 / 覆盖索引",
        "事务的 ACID 及实现", "事务隔离级别 + 默认级别", "RR 怎么（部分）解决幻读",
        "MVCC 实现原理", "redo log / undo log / binlog", "binlog 两阶段提交",
        "一条 SQL 的执行过程", "InnoDB vs MyISAM", "MySQL 有哪些锁",
        "explain 怎么看", "慢查询怎么优化",
    ],
    "JVM": [
        "JVM 内存模型（运行时数据区）", "堆和栈的区别 / 堆怎么分代",
        "程序计数器为什么线程私有", "创建对象的过程", "类加载过程",
        "双亲委派模型是什么 / 有什么用", "类加载器有哪些", "判断垃圾的方法",
        "垃圾回收算法", "垃圾回收器 / CMS vs G1", "MinorGC / MajorGC / FullGC",
        "四种引用类型", "内存泄漏 vs 内存溢出",
    ],
    "Spring": [
        "IoC / DI 是什么", "Bean 的生命周期", "Bean 的作用域",
        "三级缓存如何解决循环依赖", "AOP 原理（JDK 代理 vs CGLIB）",
        "@Transactional 事务原理", "事务的传播行为", "事务失效的场景",
        "Spring MVC 请求处理流程", "Spring Boot 自动装配原理",
        "常用注解区别（@Autowired vs @Resource 等）",
    ],
    "计算机网络": [
        "TCP 三次握手 + 为什么三次", "TCP 四次挥手 + 为什么四次 + 2MSL",
        "TCP vs UDP", "TCP 为什么可靠", "TCP 拥塞控制", "GET vs POST",
        "HTTP 常用状态码（含 502/504）", "HTTP vs HTTPS", "HTTPS 握手过程",
        "HTTP/1.1 vs 2.0", "HTTP 无状态 / Cookie 和 Session",
        "Cookie / Session / Token(JWT)", "输入 URL 到页面展示发生了什么",
        "DNS 解析流程 / TCP 还是 UDP", "HTTP 长连接 vs WebSocket",
    ],
    "Redis": [
        "Redis 为什么快", "五种数据类型及底层实现",
        "ZSet 底层 / 跳表 / 为什么不用 B+ 树", "RDB vs AOF", "内存淘汰策略",
        "过期删除策略", "缓存雪崩 / 击穿 / 穿透", "布隆过滤器原理",
        "Redis 和 MySQL 缓存一致性", "Redis 分布式锁", "大 Key / 热 Key 问题",
        "主从复制", "哨兵机制", "怎么保证 Redis 操作原子性",
    ],
    "扩展(MyBatis/MQ/分布式)": [
        "MyBatis #{} 和 ${} 的区别", "MyBatis 一级/二级缓存",
        "MQ 的作用（解耦/异步/削峰）", "MQ 如何保证消息不丢失",
        "MQ 如何保证不重复消费（幂等）", "MQ 消息积压怎么处理",
        "CAP / BASE 理论", "分布式事务方案（2PC/TCC/本地消息表）",
        "分布式 ID 生成方案（雪花算法）", "限流算法（漏桶/令牌桶等）",
    ],
    "操作系统": [
        "进程 vs 线程 vs 协程", "进程的五种状态", "进程间通信方式（IPC）",
        "线程间通信方式", "进程切换 vs 线程切换，为什么线程快", "用户态 vs 内核态",
        "死锁四条件 + 如何避免", "乐观锁 vs 悲观锁 / 自旋锁", "虚拟内存 / 地址转换",
        "程序的内存布局 / 堆栈区别", "页面置换算法",
        "IO 多路复用 / select-poll-epoll", "epoll 的 LT vs ET", "零拷贝",
    ],
    "Java基础": [
        "值传递 vs 引用传递", "装箱拆箱 + Integer 缓存（-128~127）",
        "为什么用 BigDecimal 不用 double", "重载 vs 重写", "抽象类 vs 接口",
        "final / static 的作用", "深拷贝 vs 浅拷贝 + 实现方式",
        "JVM / JDK / JRE 三者关系", "八种基本数据类型及字节数",
        "封装/继承/多态，多态的体现", "泛型是什么 + 类型擦除",
    ],
}

# 日程：每天 10 题，从明天开始
START = date(2026, 6, 10)
PER_DAY = 10
WK = ["一", "二", "三", "四", "五", "六", "日"]

def fmt_date(idx):
    d = START + timedelta(days=idx // PER_DAY)
    return f"{d.month:02d}-{d.day:02d} (周{WK[d.weekday()]})"

FONT = "Microsoft YaHei"
wb = Workbook()

# ---------- 打卡表 ----------
ws = wb.active
ws.title = "打卡表"
headers = ["板块", "序号", "建议日期", "面试问题", "掌握程度", "复习次数", "上次复习日期", "备注"]
ws.append(headers)

thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="2F5597")
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

sec_colors = {
    "集合": "FCE4D6", "并发/多线程": "FFF2CC", "MySQL": "E2EFDA", "JVM": "DDEBF7",
    "Spring": "EAD1DC", "计算机网络": "FFF2CC", "Redis": "FCE4D6",
    "扩展(MyBatis/MQ/分布式)": "D9D2E9", "操作系统": "DDEBF7", "Java基础": "E2EFDA",
}

r = 2
gidx = 0
for sec, qs in data.items():
    for i, q in enumerate(qs, 1):
        ws.cell(row=r, column=1, value=sec)
        ws.cell(row=r, column=2, value=i)
        ws.cell(row=r, column=3, value=fmt_date(gidx))
        ws.cell(row=r, column=4, value=q)
        ws.cell(row=r, column=5, value="未开始")
        ws.cell(row=r, column=6, value=0)
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=11)
            cell.border = border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if c in (2, 3, 5, 6, 7) else "left",
                                       wrap_text=(c == 4))
            if c == 1:
                cell.fill = PatternFill("solid", fgColor=sec_colors[sec])
        r += 1
        gidx += 1
last = r - 1

dv = DataValidation(type="list", formula1='"未开始,眼熟,能讲框架,能扛追问"', allow_blank=False)
ws.add_data_validation(dv)
dv.add(f"E2:E{last}")

for val, color in [("未开始", "F4CCCC"), ("眼熟", "FCE4A6"),
                   ("能讲框架", "D9EAD3"), ("能扛追问", "A9D08E")]:
    ws.conditional_formatting.add(f"E2:E{last}",
        CellIsRule(operator="equal", formula=[f'"{val}"'],
                   fill=PatternFill("solid", fgColor=color)))

for col, w in {"A": 20, "B": 6, "C": 14, "D": 44, "E": 12, "F": 10, "G": 16, "H": 24}.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{last}"

# ---------- 进度统计 ----------
ws2 = wb.create_sheet("进度统计")
ws2.append(["板块", "题数", "未开始", "眼熟", "能讲框架", "能扛追问", "完成率(能讲+能扛)"])
for c in range(1, 8):
    cell = ws2.cell(row=1, column=c)
    cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

rr = 2
for sec in data:
    ws2.cell(row=rr, column=1, value=sec)
    ws2.cell(row=rr, column=2, value=f'=COUNTIF(打卡表!$A$2:$A${last},$A{rr})')
    ws2.cell(row=rr, column=3, value=f'=COUNTIFS(打卡表!$A$2:$A${last},$A{rr},打卡表!$E$2:$E${last},"未开始")')
    ws2.cell(row=rr, column=4, value=f'=COUNTIFS(打卡表!$A$2:$A${last},$A{rr},打卡表!$E$2:$E${last},"眼熟")')
    ws2.cell(row=rr, column=5, value=f'=COUNTIFS(打卡表!$A$2:$A${last},$A{rr},打卡表!$E$2:$E${last},"能讲框架")')
    ws2.cell(row=rr, column=6, value=f'=COUNTIFS(打卡表!$A$2:$A${last},$A{rr},打卡表!$E$2:$E${last},"能扛追问")')
    ws2.cell(row=rr, column=7, value=f'=IF(B{rr}=0,0,(E{rr}+F{rr})/B{rr})')
    rr += 1
ws2.cell(row=rr, column=1, value="合计")
for col in "BCDEF":
    ws2.cell(row=rr, column="ABCDEFG".index(col) + 1, value=f'=SUM({col}2:{col}{rr-1})')
ws2.cell(row=rr, column=7, value=f'=IF(B{rr}=0,0,(E{rr}+F{rr})/B{rr})')

for c in range(1, 8):
    for row in range(1, rr + 1):
        cell = ws2.cell(row=row, column=c)
        cell.border = border
        cell.font = Font(name=FONT, size=11, bold=(row in (1, rr)))
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")
        if c == 7 and row >= 2:
            cell.number_format = "0%"
total_fill = PatternFill("solid", fgColor="FFF2CC")
for c in range(1, 8):
    ws2.cell(row=rr, column=c).fill = total_fill
ws2.cell(row=1, column=c)
for col, w in {"A": 22, "B": 8, "C": 9, "D": 9, "E": 11, "F": 11, "G": 18}.items():
    ws2.column_dimensions[col].width = w
ws2.freeze_panes = "A2"

# ---------- 用法说明 ----------
ws3 = wb.create_sheet("用法说明")
notes = [
    ["秋招后端必背清单 · 打卡表", ""],
    ["", ""],
    ["怎么用", ""],
    ["1", "「建议日期」列已按优先级排好计划：每天 10 题，从 6/10 开始，约两周过完第一遍。跟着日期打卡即可。"],
    ["2", "每过一道题，更新「掌握程度」列（点单元格有下拉），四档：未开始→眼熟→能讲框架→能扛追问。"],
    ["3", "目标是把每题推到「能扛追问」。每复习一次「复习次数」+1，同一题建议滚动 3 遍以上。"],
    ["4", "「上次复习日期」配合间隔重复：隔 1 天、3 天、7 天各回顾一次。"],
    ["5", "「进度统计」自动汇总各板块完成率，一眼看到哪块还薄弱。"],
    ["", ""],
    ["计划说明", "第一遍最慢（理解+自己讲一遍）；二刷只默写卡壳题，速度翻几倍。别等背完再面试，前几场面试就是最好的复习。"],
    ["复习优先级", "集合≈并发＞MySQL≈JVM＞Spring＞网络＞Redis＞扩展＞操作系统＞Java基础"],
    ["配套文档", "《秋招后端八股必背清单.md》= 分档清单；《秋招必背答案.md》= 每题口述答案"],
    ["共计", "132 道 ⭐ 必背题"],
]
for row in notes:
    ws3.append(row)
ws3["A1"].font = Font(name=FONT, bold=True, size=14, color="2F5597")
for row in range(3, len(notes) + 1):
    ws3.cell(row=row, column=1).font = Font(name=FONT, bold=True, size=11)
    ws3.cell(row=row, column=2).font = Font(name=FONT, size=11)
    ws3.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="center")
ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 78

wb.save("秋招必背打卡表.xlsx")
print("saved, total =", last - 1)
